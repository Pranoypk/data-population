%%time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import re
 
# Load the first dataset (initially provided data)
df1 = pd.read_excel(r"C:\Users\pkayyala\Desktop\December\Sandvik_product_file_final.xlsx",
    sheet_name="pimattribute",
    dtype={"mfg_part": str}  # Ensure mfg_part is read as a string
)
df1["mfg_part"] = df1["mfg_part"].astype(str)
df1["Mill_Diameter"] = df1["Cutting_Diameter"]
df1["Diameter"] = df1["Cutting_Diameter"]
df1["Size"] = df1["Cutting_Diameter"]
df1["Drilling_diameter_Inch_"] = df1["Cutting_Diameter"]
df1["Drill_Size"] = df1["Cutting_Diameter"]
df1["Toolholder_Material"] = df1["Material"]
df1["Surface_Material"] = df1["Material"]
df1["Twist_Drill_Material"] = df1["Material"]
df1["Drill_Material"] = df1["Material"]
df1["Face_effective_cutting_edge_count"] = df1["Flute"]
df1["Number_of_Cutter_Inserts"] = df1["Flute"]
#df1["Grade"] = df1["Manufacturer_s_Grade"]
df1["Spiral_Direction"]= df1["Cutting_Direction"]
df1["Flute_Direction"]= df1["Cutting_Direction"]
df1["Thread_Direction"]= df1["Cutting_Direction"]
df1["Insert_Hand"]= df1["Cutting_Direction"]
df1["Insert_Hand"] = df1["Insert_Hand"].str.replace("LEFTHAND", "LEFT_HAND")
df1["Thickness_Inch_"]=df1["Insert_thickness_Inch_"]
df1["Maximum_Drill_Size"]= df1["Maximum_Cutting_Diameter"]
df1["Minimum_Drill_Size"]= df1["Minimum_Cutting_Diameter"]
df1["Minimum_Drill_Diameter"]= df1["Minimum_Cutting_Diameter"]
df1["Minimum_Drill_Bit_Size"]= df1["Minimum_Cutting_Diameter"]
df1["Number_of_Teeth"]= df1["Flute"]
#df1["Protruding_length_Inch_"] = df1["Projection"]
df1["Nose_Diameter"]= df1["Body_Diameter"]
df1["Drill_Point_Angle"]=df1["Point_Angle"]
df1["Point_angle_1st_Step"]=df1["Point_Angle"]
df1["Length_chip_flute"]=df1["Flute_Length"]
#df1["Maximum_Depth_of_Cut"]=df1["Depth_Of_Cut"]



# Load the second dataset (newly provided data)
df2 = pd.read_excel(r"C:\Users\pkayyala\Desktop\September\master_data.xlsx")
 
# Colors for highlighting (not used in the script)
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
 
# Get unique family codes from df1
Family_Names = df1['Family_Name'].unique()
 
# Directory for saving template files
output_dir = r"C:\Users\pkayyala\Desktop\December\sandvik_update_2"
# Ensure the directory exists
os.makedirs(output_dir, exist_ok=True)

# Function to sanitize file names
def sanitize_filename(filename):
    return re.sub(r'[<>:"/\\|?*]', '_', filename)

# Initialize a list to store file paths and missing value percentages
file_info = []
missing_mfg_part_families = []  # List to store Family Names missing 'mfg_part'

# Loop through each family code and create template files
for Family_Name in Family_Names:
    # Filter df1 based on the current family code
    filtered_df1 = df1[df1['Family_Name'] == Family_Name]
    # Filter df2 based on the current family code (if applicable)
    filtered_df2 = df2[df2['Family_Name'] == Family_Name] if 'Family_Name' in df2.columns else df2
    
    # Create an empty DataFrame with headers from filtered_df2
    template_df = pd.DataFrame(columns=filtered_df2['Attribute_Code'])
    
    # Check if 'mfg_part' is missing in the columns
    if "mfg_part" not in template_df.columns:
        missing_mfg_part_families.append(Family_Name)
        print(f"'mfg_part' is missing in template for Family_Name: {Family_Name}")

    # Populate the template_df with values from filtered_df1
    for col in template_df.columns:
        if col in filtered_df1.columns:
            template_df[col] = filtered_df1[col].values

    # Add columns with default values if necessary
    if "mfg_part" not in template_df.columns:
        template_df["mfg_part"] = ""
    
    template_df["Enable_In_EGC_Supply"] = 0
    template_df["mfg_part"] = template_df["mfg_part"].astype(str)
    template_df["enabled"] = 1
    template_df["special_item"] = 0

    # Calculate the missing value percentage for each column
    missing_percentages = template_df.isna().mean() * 100
    # Create the file path with sanitized file name
    sanitized_family_name = sanitize_filename(Family_Name)
    file_path = os.path.join(output_dir, f'{sanitized_family_name}_template.xlsx')
    # Save the template DataFrame to Excel
    template_df.to_excel(file_path, index=False)
    # Store the file path and missing value percentages
    file_info.append({
        'file_path': file_path,
        'missing_percentages': missing_percentages.to_dict()  # Convert to dictionary for better readability
    })

# Print or save the list of Family Names without 'mfg_part'
print("Templates missing 'mfg_part':", missing_mfg_part_families)

file_info
